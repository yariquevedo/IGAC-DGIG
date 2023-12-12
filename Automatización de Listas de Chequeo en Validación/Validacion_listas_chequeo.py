# -*- coding: utf-8 -*-
# ---------------------------------------------------------------------------
# Direccion de Gestion de Información Geografica
# Created on: 2023-08-09
# Created by: Kelly Villamil - Diego Rugeles (Supervisor Desarrollo DGIG)
# # Usage: Validacion productos cartográficos
# Description:
# ---------------------------------------------------------------------------
#Librerias
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import shutil 
import os
import arcpy
from arcpy import env
from arcpy.sa import *
#Parametros
lista_chequeo = arcpy.GetParameterAsText(0)
datos_fuente = arcpy.GetParameterAsText(1)
salida = arcpy.GetParameterAsText(2)
def diligenciar(lista_chequeo, datos_fuente,salida):
    arcpy.env.overwriteOutput = True
    arcpy.management.CreateFolder(salida, 'temp')
    temp = os.path.join(str(salida),'temp')
    wb_2 = load_workbook(datos_fuente)
    df = wb_2["Hoja1"]
    i = 0 
    
    for row in df.iter_rows(min_row=3, min_col=1):
        lc = shutil.copy2(lista_chequeo,str(temp+ "\\" + "LC_orto" + str(i) + ".xlsx"))
        wb_1 = load_workbook(lc)
        
        hoja_mdt = wb_1["Modelo_Digital_Terreno(MDT)"]
        hoja_mdt["D6"] = row[0].value #proyecto/entidad territorial
        hoja_mdt["D9"] = row[1].value #ubicacion producto
        hoja_mdt["J7"] = row[2].value #escala
        hoja_mdt["K9"] = row[3].value #año
        hoja_mdt["L9"] = row[4].value #mes
        hoja_mdt["M9"] = row[5].value #dia
        hoja_mdt["J9"] = row[7].value #grilla
        hoja_mdt["J11"] = row[10].value #area
        hoja_orto = wb_1["Ortofoto_Ortoimagen_Mosaico"]
        hoja_orto["D6"] = row[0].value #proyecto/entidad territorial
        hoja_orto["D9"] = row[1].value #ubicacion producto
        hoja_orto["J7"] = row[11].value #escala
        hoja_orto["K9"] = row[12].value #año
        hoja_orto["L9"] = row[13].value #mes
        hoja_orto["M9"] = row[14].value #dia
        hoja_orto["J9"] = row[15].value #GSD
        hoja_orto["J11"] = row[19].value #area
        hoja_gdb = wb_1["Base_de_datos_cartográfica"]
        hoja_gdb["D6"] = row[0].value #proyecto/entidad territorial
        hoja_gdb["D9"] = row[1].value #ubicacion producto
        hoja_gdb["J7"] = row[20].value #escala
        hoja_gdb["K9"] = row[21].value #año
        hoja_gdb["L9"] = row[22].value #mes
        hoja_gdb["M9"] = row[23].value #dia
        hoja_gdb["J11"] = row[28].value #area
        
        
        if str(row[8].value) == 'None':
            if str(row[17].value) == 'None':
                ent = str(row[26].value)
            else:
                ent = str(row[17].value)
        else:
            ent = str(row[8].value)
        if str(row[9].value) == 'None':
            if str(row[18].value) == 'None':
                esc = str(row[27].value)
            else:
                esc = str(row[18].value)
        else:
            esc = str(row[9].value)
        wb_1.save(os.path.join(salida,str("Verificacion_Inicial_"+ent +"_"+esc+".xlsx")))
        os.remove(lc)
        i = i+1
    arcpy.management.Delete(os.path.join(str(salida),'temp'))
    
if __name__ == '__main__':
    arcpy.AddMessage("Generando los archivos de chequeo")
    diligenciar(lista_chequeo, datos_fuente,salida)
    contenido = os.listdir(salida)
    n = len(contenido)
    arcpy.AddMessage("Se generaron " +str(n) + " archivos en la ruta de salida indicada")
    arcpy.AddMessage("Finalizado")
